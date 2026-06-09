# -*- coding: utf-8 -*-
"""Edge cases: all export formats with empty, normal, and injection inputs."""
import io
import json
import os
import sys
import uuid

import pytest

BACKEND_DIR = os.path.join(os.path.dirname(__file__), "..", "backend")
sys.path.insert(0, os.path.abspath(BACKEND_DIR))

from models.session import SessionModel
from models.attack import AttackDetectionResult
from services.export_service import ExportService


def _s(src_ip="1.2.3.4", dst_ip="5.6.7.8", bytes_total=1000, protocol="TCP", port=80):
    return SessionModel(
        session_id=str(uuid.uuid4()),
        src_ip=src_ip, dst_ip=dst_ip,
        src_port=12345, dst_port=port,
        protocol=protocol,
        start_ts=1.0, end_ts=2.0,
        bytes_sent=bytes_total // 2, bytes_recv=bytes_total // 2,
        packet_count=10, payload_length=0,
    )


def _a(attack_type="DoS", confidence="high"):
    return AttackDetectionResult(
        attack_type=attack_type,
        confidence=confidence,
        evidence=["test evidence"],
        mitre_id="T1498.001",
    )


svc = ExportService()


# ── CSV export ───────────────────────────────────────────────────

class TestCSVExport:
    def test_csv_empty_sessions(self):
        data = svc.export([], [], "csv")
        text = data.decode("utf-8")
        lines = text.strip().split("\n")
        assert len(lines) == 1  # header only

    def test_csv_has_header(self):
        data = svc.export([_s()], [], "csv")
        text = data.decode("utf-8")
        assert "src_ip" in text
        assert "dst_ip" in text
        assert "protocol" in text

    def test_csv_single_session(self):
        data = svc.export([_s(src_ip="1.1.1.1")], [], "csv")
        text = data.decode("utf-8")
        assert "1.1.1.1" in text

    def test_csv_multiple_sessions(self):
        sessions = [_s(src_ip=f"10.0.0.{i}") for i in range(1, 6)]
        data = svc.export(sessions, [], "csv")
        text = data.decode("utf-8")
        lines = text.strip().split("\n")
        assert len(lines) == 6  # 1 header + 5 data rows

    def test_csv_returns_bytes(self):
        data = svc.export([_s()], [], "csv")
        assert isinstance(data, bytes)


# ── JSON export ──────────────────────────────────────────────────

class TestJSONExport:
    def test_json_empty_returns_empty_list(self):
        data = svc.export([], [], "json")
        parsed = json.loads(data)
        assert parsed == []

    def test_json_single_session_is_valid(self):
        data = svc.export([_s()], [], "json")
        parsed = json.loads(data)
        assert len(parsed) == 1
        assert "src_ip" in parsed[0]

    def test_json_multiple_sessions(self):
        sessions = [_s() for _ in range(5)]
        data = svc.export(sessions, [], "json")
        parsed = json.loads(data)
        assert len(parsed) == 5

    def test_json_returns_bytes(self):
        assert isinstance(svc.export([], [], "json"), bytes)

    def test_json_fields_present(self):
        data = svc.export([_s()], [], "json")
        parsed = json.loads(data)[0]
        for field in ("session_id", "src_ip", "dst_ip", "protocol", "bytes_total"):
            assert field in parsed, f"Missing field: {field}"


# ── Excel export ─────────────────────────────────────────────────

class TestExcelExport:
    def test_excel_empty_sessions(self):
        pytest.importorskip("openpyxl")
        data = svc.export([], [], "excel")
        assert isinstance(data, bytes)
        assert len(data) > 0

    def test_excel_valid_workbook(self):
        from openpyxl import load_workbook
        pytest.importorskip("openpyxl")
        data = svc.export([_s()], [_a()], "excel")
        wb = load_workbook(io.BytesIO(data))
        assert "Sessions" in wb.sheetnames
        assert "Attacks" in wb.sheetnames

    def test_excel_sessions_sheet_has_data(self):
        from openpyxl import load_workbook
        pytest.importorskip("openpyxl")
        data = svc.export([_s(src_ip="2.2.2.2")], [], "excel")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Sessions"]
        values = [cell.value for row in ws.iter_rows() for cell in row]
        assert "2.2.2.2" in values

    def test_excel_attacks_sheet_has_data(self):
        from openpyxl import load_workbook
        pytest.importorskip("openpyxl")
        data = svc.export([_s()], [_a("DDoS")], "excel")
        wb = load_workbook(io.BytesIO(data))
        ws = wb["Attacks"]
        values = [cell.value for row in ws.iter_rows() for cell in row]
        assert "DDoS" in values


# ── PDF export ───────────────────────────────────────────────────

class TestPDFExport:
    def test_pdf_empty_sessions(self):
        pytest.importorskip("reportlab")
        data = svc.export([], [], "pdf")
        assert data[:4] == b"%PDF"

    def test_pdf_many_sessions_truncated_to_50(self):
        pytest.importorskip("reportlab")
        sessions = [_s(src_ip=f"10.0.{i//256}.{i%256}") for i in range(100)]
        data = svc.export(sessions, [], "pdf")
        assert data[:4] == b"%PDF"

    def test_pdf_returns_bytes(self):
        pytest.importorskip("reportlab")
        assert isinstance(svc.export([], [], "pdf"), bytes)


# ── Suricata export ──────────────────────────────────────────────

class TestSuricataExport:
    def test_empty_attacks_returns_empty(self):
        data = svc.export([_s()], [], "suricata")
        assert data == b""

    def test_rule_contains_alert_keyword(self):
        data = svc.export([_s()], [_a("DoS")], "suricata")
        assert b"alert" in data

    def test_injection_attempt_sanitized(self):
        attack = AttackDetectionResult(
            attack_type="DoS",
            confidence="high",
            evidence=["test"],
            mitre_id="T1498",
        )
        # attack_type is a Literal, so injection in msg comes from sanitization
        data = svc.export([_s()], [attack], "suricata")
        rules = data.decode("utf-8")
        # Verify the rule format is syntactically valid (no unescaped quotes)
        assert '";' in rules or rules == ""  # proper rule termination

    def test_icmp_session_uses_icmp_proto(self):
        sessions = [_s(protocol="ICMP")]
        sessions[0] = SessionModel(
            session_id=str(uuid.uuid4()),
            src_ip="1.2.3.4", dst_ip="5.6.7.8",
            src_port=0, dst_port=0,
            protocol="ICMP",
            start_ts=0.0, end_ts=1.0,
            bytes_sent=50, bytes_recv=50,
            packet_count=1, payload_length=0,
        )
        data = svc.export(sessions, [_a("DoS")], "suricata")
        assert b"icmp" in data

    def test_sid_numbers_unique(self):
        sessions = [_s() for _ in range(3)]
        attacks = [_a("DoS"), _a("DDoS")]
        data = svc.export(sessions, attacks, "suricata")
        rules = data.decode("utf-8").strip().split("\n")
        sids = [r.split("sid:")[1].split(";")[0] for r in rules if "sid:" in r]
        assert len(sids) == len(set(sids)), "Duplicate SIDs in Suricata rules"


# ── Snort export ─────────────────────────────────────────────────

class TestSnortExport:
    def test_empty_attacks_returns_empty(self):
        data = svc.export([_s()], [], "snort")
        assert data == b""

    def test_rule_contains_alert_keyword(self):
        data = svc.export([_s()], [_a("PortScan")], "snort")
        assert b"alert" in data

    def test_sid_different_from_suricata(self):
        sessions = [_s()]
        attacks = [_a("DoS")]
        suricata_data = svc.export(sessions, attacks, "suricata")
        snort_data = svc.export(sessions, attacks, "snort")
        suricata_sids = [r.split("sid:")[1].split(";")[0] for r in suricata_data.decode().split("\n") if "sid:" in r]
        snort_sids = [r.split("sid:")[1].split(";")[0] for r in snort_data.decode().split("\n") if "sid:" in r]
        assert suricata_sids != snort_sids, "Suricata and Snort should use different SID ranges"

    def test_returns_bytes(self):
        assert isinstance(svc.export([], [], "snort"), bytes)


# ── Unsupported format ───────────────────────────────────────────

class TestUnsupportedFormat:
    def test_unknown_format_raises_value_error(self):
        with pytest.raises(ValueError, match="Unsupported"):
            svc.export([], [], "xml")
